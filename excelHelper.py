#Le format de fichier Excel doit être le suivant:
#Colonne B:Time
#Colonne C:Latitude
#Colonne D:Longitude
#Colonne E:Altitude
from .pointRoute import *
#À modifier si la structure fichier change.
DEBUTLIGNE = 9
COLLATITUDE=3
COLLONGITUDE=4
COLALTITUDE=5
COLDATE=2

#Vérification si le module existe.
import openpyxl
class ExcelHelper:
    def __init__(self, cheminAcces, signal):
        self.path = cheminAcces
        self._workbook = None
        self._listeFeuillets = None
        self._listePoints = []
        self._signal = signal
        self._ouverture()

    def log(self, message:str) -> None:
        if self._signal:
            self._signal.signal.emit(message)
        print(message)

    def _ouverture(self):
        self.log("Ouverture du fichier Excel")
        try:
            self._workbook = openpyxl.load_workbook(self.path)
            self._listeFeuillets = self._workbook.sheetnames  # Recherche des feuillets
            self.log("Les feuillets du fichier Excel: ", self._listeFeuillets)
        except Exception as e:
            self.log(f"Problèmes {__name__}: {e}")
        finally:
            self._workbook.close()

    def extractionPointFeuillet(self)->list:
        for feuillet in self._listeFeuillets:
            _ligne = DEBUTLIGNE
            self.log("Traitement du feuillet {}".format(feuillet.title()))
            feuillet = self._workbook[feuillet]
            data = feuillet.cell(_ligne, COLDATE).value
            while data != None:
                self.log(f"Traitement de la ligne {_ligne}")
                #Création d'un point et ajout dans la liste
                self._listePoints.append(Point(feuillet.cell(_ligne, COLLATITUDE).value, feuillet.cell(_ligne, COLLONGITUDE).value, feuillet.cell(_ligne, COLALTITUDE).value, feuillet.cell(_ligne, COLDATE).value,  feuillet.title))
                _ligne += 1
                data = feuillet.cell(_ligne, COLDATE).value
            #print(self._listePoints)
        self._fermeture()
        return self._listePoints

    def _fermeture(self):
        self.log("Fermeture du fichier Excel")
        if self._workbook != None:
            self._workbook.close()
